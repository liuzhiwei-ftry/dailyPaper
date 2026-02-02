import os
from datetime import datetime
from openpyxl import Workbook
from PySide6.QtWidgets import QFileDialog

class ExcelExporter:
    """Excel导出工具"""
    @staticmethod
    def export_to_excel(content: str, work_content: str = "") -> bool:
        """导出日报到Excel"""
        if not content.strip():
            return False

        # 选择保存路径
        save_path, _ = QFileDialog.getSaveFileName(
            None,
            "导出Excel",
            f"日报_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
            "Excel文件 (*.xlsx)"
        )
        if not save_path:
            return False

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "AI日报"

        # 写入内容
        ws["A1"] = "生成时间"
        ws["B1"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A2"] = "当日工作内容"
        ws["B2"] = work_content
        ws["A3"] = "生成日报内容"
        # 按行拆分内容
        content_lines = content.split("\n")
        for idx, line in enumerate(content_lines, start=3):
            ws[f"B{idx}"] = line

        # 保存文件
        try:
            wb.save(save_path)
            return True
        except Exception as e:
            print(f"Excel导出失败：{e}")
            return False